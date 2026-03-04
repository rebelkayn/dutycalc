"""
build_hts.py — Download and convert USITC HTS Excel into hts_data.json

Fully automatic — no manual download needed.

Tries each revision in descending order (latest first):
  https://www.usitc.gov/sites/default/files/tata/hts/bychapter/hts_2026_revision_4_xls.xlsx
  https://www.usitc.gov/sites/default/files/tata/hts/bychapter/hts_2026_revision_3_xls.xlsx
  ...

When a new revision is released, the script detects it automatically.
Stored revision number is saved to data/us_version.txt.

Usage:
    python build_hts.py           # Auto-download latest, skip if up to date
    python build_hts.py --force   # Force re-download and rebuild

Output: data/hts_data.json — compact JSON loaded by app.py at startup
"""

import json
import os
import re
import sys
import urllib.request
import urllib.error

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl

DATA_DIR     = os.path.join(os.path.dirname(__file__), "data")
EXCEL_PATH   = os.path.join(DATA_DIR, "hts_2026.xlsx")
OUTPUT_PATH  = os.path.join(DATA_DIR, "hts_data.json")
VERSION_PATH = os.path.join(DATA_DIR, "us_version.txt")

YEAR = 2026
# Try revisions from high to low — stops at first successful download
MAX_REVISION = 9
MIN_REVISION = 1

def revision_url(year: int, rev: int) -> str:
    return (
        f"https://www.usitc.gov/sites/default/files/tata/hts/bychapter/"
        f"hts_{year}_revision_{rev}_xls.xlsx"
    )

def load_stored_version() -> str:
    if os.path.exists(VERSION_PATH):
        with open(VERSION_PATH) as f:
            return f.read().strip()
    return ""

def save_version(v: str):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(VERSION_PATH, "w") as f:
        f.write(v)

def find_and_download(force=False) -> tuple[str, int]:
    """
    Try revisions from MAX down to MIN. Returns (local_path, revision).
    Skips download if stored version matches latest found, unless force=True.
    """
    stored = load_stored_version()

    for rev in range(MAX_REVISION, MIN_REVISION - 1, -1):
        url = revision_url(YEAR, rev)
        version_str = f"{YEAR}_rev{rev}"

        # HEAD check — see if this URL exists without downloading
        try:
            req = urllib.request.Request(url, method="HEAD",
                                         headers={"User-Agent": "iDuties/1.0"})
            with urllib.request.urlopen(req, timeout=10) as r:
                size = int(r.headers.get("Content-Length", 0))
                if size < 10000:  # Too small — not a valid Excel
                    continue
        except urllib.error.HTTPError as e:
            if e.code == 404:
                continue  # This revision doesn't exist yet
            continue
        except Exception:
            continue

        # Found the latest revision
        print(f"  Latest revision : {YEAR} Rev {rev}")
        print(f"  Stored revision : {stored or '(none)'}")

        if not force and version_str == stored and os.path.exists(OUTPUT_PATH):
            print(f"✓  Already up to date ({version_str})")
            print(f"   Use --force to rebuild anyway")
            return None, rev  # Signal: no rebuild needed

        # Download it
        print(f"\n  Downloading HTS {YEAR} Revision {rev}...")
        print(f"  URL: {url}")
        os.makedirs(DATA_DIR, exist_ok=True)

        def progress(count, block, total):
            pct = min(100, count * block * 100 // total) if total else 0
            print(f"\r  Progress: {pct}%", end="", flush=True)

        urllib.request.urlretrieve(url, EXCEL_PATH, reporthook=progress)
        print(f"\r  Downloaded {os.path.getsize(EXCEL_PATH)//1024} KB          ")

        return EXCEL_PATH, rev

    print(f"❌  No valid HTS revision found for {YEAR} (tried Rev {MIN_REVISION}–{MAX_REVISION})")
    print(f"   Check: https://www.usitc.gov/harmonized_tariff_information")
    sys.exit(1)

# Section 301 China surcharges by HTS chapter/heading
# Source: USTR Section 301 Lists 1-4A (as of 2026)
# List 1 (25%): machinery, electronics components, industrial goods
# List 2 (25%): semiconductors, chemicals, plastics
# List 3 (25%): consumer goods, furniture, apparel
# List 4A (7.5%): consumer goods (reduced from 15% in Phase 1 deal)
CN_301_RATES = {
    # 7.5% (List 4A) — consumer goods
    **{str(c): 7.5 for c in [
        61, 62, 63, 64,   # apparel, footwear, textiles
        42, 43,           # bags, leather
        95, 96,           # toys, misc
        90, 91,           # instruments, watches
        39, 40,           # plastics, rubber
        44, 45, 46,       # wood products
        47, 48, 49,       # paper
        69, 70,           # ceramics, glass
        94,               # furniture
    ]},
    # 25% (Lists 1-3) — industrial, electronics, metals
    **{str(c): 25.0 for c in [
        72, 73, 74, 75, 76, 78, 79, 80, 81,  # metals
        82, 83,           # tools, hardware
        84,               # machinery
        85,               # electrical/electronics
        86, 87, 88, 89,   # vehicles, aircraft, ships
        68,               # stone, cement, asbestos
    ]},
    # 0% — excluded categories
    **{str(c): 0.0 for c in [
        1, 2, 3, 4, 5,    # live animals, meat, fish, dairy
        6, 7, 8, 9, 10,   # plants, vegetables, fruit, coffee, cereals
        11, 12, 13, 14,   # milling, oilseeds, gums
        15,               # fats and oils
        16, 17, 18, 19,   # prepared foods
        20, 21, 22, 23,   # beverages, vinegar
        24,               # tobacco
        25, 26, 27,       # minerals, ores, fuels
        28, 29, 30,       # chemicals, pharma
        31, 32, 33, 34, 35, 36, 37, 38,  # fertilizers, cosmetics
        93,               # arms and ammunition
        98, 99,           # special classification provisions
    ]},
}

def get_cn_301(hts_code: str) -> float:
    """Look up China Section 301 surcharge for an HTS code."""
    # Strip dots and get chapter (first 2 digits)
    clean = hts_code.replace(".", "").replace(" ", "")
    if len(clean) < 2:
        return 0.0
    chapter = clean[:2]
    return CN_301_RATES.get(chapter, 0.0)

def parse_rate(rate_str) -> float:
    """Parse a duty rate string like '16.5%', 'Free', '7.5¢/kg + 6%' into a float percentage."""
    if not rate_str:
        return 0.0
    s = str(rate_str).strip().lower()
    if s in ("", "free", "0", "0.0", "none", "-"):
        return 0.0
    # Extract first percentage found
    m = re.search(r"(\d+\.?\d*)\s*%", s)
    if m:
        return float(m.group(1))
    # Pure number (already a percentage)
    m = re.search(r"^(\d+\.?\d*)$", s)
    if m:
        return float(m.group(1))
    return 0.0

def normalize_code(raw) -> str:
    """Normalize HTS code to XXXX.XX.XX format."""
    if not raw:
        return ""
    s = str(raw).strip().replace(" ", "").replace("\u2013", "").replace("\u2014", "")
    digits = re.sub(r"[^\d]", "", s)
    if len(digits) < 6:
        return ""
    # Format as XXXX.XX.XX (10-digit) or XXXX.XX (6-digit)
    if len(digits) >= 8:
        return f"{digits[:4]}.{digits[4:6]}.{digits[6:8]}"
    elif len(digits) >= 6:
        return f"{digits[:4]}.{digits[4:6]}"
    return ""

def find_columns(ws):
    """Find which columns contain HTS code, description, and general rate."""
    header_row = None
    col_map = {"code": None, "desc": None, "rate": None}

    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        row_str = [str(c).lower() if c else "" for c in row]
        row_joined = " ".join(row_str)
        if any(kw in row_joined for kw in ["hts", "tariff", "heading", "subheading"]):
            header_row = list(row)
            break

    if not header_row:
        # Guess: code in col 0, desc in col 1, rate somewhere around col 3-6
        col_map = {"code": 0, "desc": 1, "rate": 4}
        return col_map

    for i, cell in enumerate(header_row):
        if not cell:
            continue
        val = str(cell).lower().strip()
        if col_map["code"] is None and any(kw in val for kw in ["hts", "heading", "subheading", "number"]):
            col_map["code"] = i
        if col_map["desc"] is None and any(kw in val for kw in ["description", "article", "commodity"]):
            col_map["desc"] = i
        if col_map["rate"] is None and any(kw in val for kw in ["general", "rate", "mfn", "normal"]):
            col_map["rate"] = i

    # Fallback defaults
    if col_map["code"] is None: col_map["code"] = 0
    if col_map["desc"] is None: col_map["desc"] = 1
    if col_map["rate"] is None: col_map["rate"] = 4

    return col_map

def build_json(force=False):
    print("🇺🇸  USITC HTS Builder")
    print("─" * 40)

    excel_path, rev = find_and_download(force=force)
    if excel_path is None:
        return False  # Already up to date

    print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Try to find the right sheet (usually named "HTS" or first sheet)
    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if any(kw in name.lower() for kw in ["hts", "tariff", "schedule"]):
            sheet_name = name
            break
    ws = wb[sheet_name]
    print(f"  Using sheet: '{sheet_name}'")

    col_map = find_columns(ws)
    print(f"  Columns — code:{col_map['code']} desc:{col_map['desc']} rate:{col_map['rate']}")

    hts_data = {}
    skipped = 0
    processed = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            code_raw = row[col_map["code"]] if col_map["code"] < len(row) else None
            desc_raw = row[col_map["desc"]] if col_map["desc"] < len(row) else None
            rate_raw = row[col_map["rate"]] if col_map["rate"] < len(row) else None

            code = normalize_code(code_raw)
            if not code or len(code) < 7:
                skipped += 1
                continue

            desc = str(desc_raw).strip() if desc_raw else ""
            if not desc or desc.lower() in ("none", "nan", ""):
                skipped += 1
                continue

            rate = parse_rate(rate_raw)
            cn301 = get_cn_301(code)
            chapter = int(code[:2]) if code[:2].isdigit() else 0

            hts_data[code] = {
                "d": desc[:120],
                "r": rate,
                "c": cn301,
                "ch": chapter,
            }
            processed += 1

            if processed % 1000 == 0:
                print(f"  Processed {processed} codes...")

        except Exception:
            skipped += 1
            continue

    wb.close()

    print(f"\n✓  Parsed {processed} HTS codes ({skipped} rows skipped)")

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(hts_data, f, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"✓  Written to {OUTPUT_PATH}  ({size_kb:.0f} KB)")

    version_str = f"{YEAR}_rev{rev}"
    save_version(version_str)
    print(f"✓  Version saved: {version_str}")

    print(f"\nNext steps:")
    print(f"  git add data/hts_data.json data/us_version.txt")
    print(f"  git commit -m 'US HTS update: {version_str} ({processed} codes)'")
    print(f"  git push")
    return True

if __name__ == "__main__":
    force = "--force" in sys.argv
    build_json(force=force)
