"""
build_eu.py — Build EU TARIC tariff database from European Commission

Source: EU Taxation and Customs Union — TARIC raw data
Data:   Combined Nomenclature (CN) + Common Customs Tariff (CCT) 2026
Format: Excel file published annually by the European Commission

The EU publishes two useful sources:
  1. TARIC Excel bulk download (via data.europa.eu)
  2. EUR-Lex Official Journal — annual CN regulation (PDF/XML)

Best programmatic source: data.europa.eu SPARQL/API
Fallback: CIRCABC Excel (requires EU login — not automatable)
Practical: Download from data.europa.eu or use TARIC XML daily snapshot

Auto-update: Checks EU Open Data Portal for dataset modification date.

Usage:
    # Option A — automatic download attempt:
    python builders/build_eu.py

    # Option B — manual Excel path:
    python builders/build_eu.py --excel path/to/taric_2026.xlsx

Output: data/eu_tariff.json
"""

import json
import csv
import os
import re
import sys
import urllib.request
import urllib.error
import io
import zipfile

DATA_DIR     = os.path.join(os.path.dirname(__file__), "..", "data")
OUTPUT_PATH  = os.path.join(DATA_DIR, "eu_tariff.json")
VERSION_PATH = os.path.join(DATA_DIR, "eu_version.txt")

# ── EU Data Sources ────────────────────────────────────────────────────────────
# data.europa.eu hosts the TARIC dataset as open data
# The CSV export URL — check data.europa.eu/data/datasets/eu-customs-tariff-taric
# for the latest resource URL. Updated annually around November.
TARIC_CSV_URL = (
    "https://data.europa.eu/api/hub/store/data/"
    "eu-customs-tariff-taric/taric_publication_20260101.csv"
)
# Fallback: WTO Tariff Download Facility for EU (requires form submission — manual)
# Best fallback: OpenTariff / CIRCABC Excel

# TARIC consultation API (per-code lookup — slow but always current)
TARIC_API_URL = "https://ec.europa.eu/taxation_customs/dds2/taric/measures.jsp"

META_URL = "https://data.europa.eu/api/hub/search/datasets/eu-customs-tariff-taric"

# ── Rate parser ────────────────────────────────────────────────────────────────
def parse_rate(s) -> float:
    if not s:
        return 0.0
    s = str(s).strip().lower()
    if s in ("", "free", "0", "0.0", "-", "n/a", "exempt", "—"):
        return 0.0
    m = re.search(r"(\d+\.?\d*)\s*%", s)
    if m:
        return float(m.group(1))
    m = re.match(r"^(\d+\.?\d*)$", s)
    if m:
        return float(m.group(1))
    return 0.0

def normalize_code(raw) -> str:
    if not raw:
        return ""
    digits = re.sub(r"[^\d]", "", str(raw).strip())
    if len(digits) < 6:
        return ""
    return digits[:10]

def check_version() -> str:
    try:
        req = urllib.request.Request(META_URL, headers={"Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            meta = json.loads(r.read())
            modified = meta.get("result", {}).get("modified", "")
            return modified[:10] if modified else "unknown"
    except Exception:
        return "unknown"

def load_stored_version() -> str:
    if os.path.exists(VERSION_PATH):
        with open(VERSION_PATH) as f:
            return f.read().strip()
    return ""

def save_version(v: str):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(VERSION_PATH, "w") as f:
        f.write(v)

def parse_csv(raw_text: str) -> dict:
    """Parse TARIC CSV into tariff dict."""
    reader = csv.DictReader(io.StringIO(raw_text))
    fieldnames = reader.fieldnames or []
    print(f"  Columns: {', '.join(str(f) for f in fieldnames[:10])}")

    # TARIC CSV typical columns:
    # NOMENCLATURE, GOODS_DESCRIPTION, MEASURE_TYPE, DUTY_AMOUNT, START_DATE, ...
    code_col = next((c for c in fieldnames if any(k in str(c).lower() for k in
                     ["nomenclature", "cn_code", "commodity", "tariff_code"])), None)
    desc_col = next((c for c in fieldnames if any(k in str(c).lower() for k in
                     ["description", "goods_desc"])), None)
    rate_col = next((c for c in fieldnames if any(k in str(c).lower() for k in
                     ["duty_amount", "duty_rate", "rate"])), None)
    type_col = next((c for c in fieldnames if "measure_type" in str(c).lower()), None)

    if not code_col:
        code_col = fieldnames[0] if fieldnames else None
    if not rate_col:
        rate_col = fieldnames[3] if len(fieldnames) > 3 else None

    print(f"  Using — code:'{code_col}'  desc:'{desc_col}'  rate:'{rate_col}'")

    tariff = {}
    skipped = 0
    processed = 0

    for row in reader:
        try:
            # Only keep MFN (measure type 103 = third country duty)
            if type_col:
                mt = str(row.get(type_col, "")).strip()
                if mt and mt not in ("103", "MFN", "Third country duty", ""):
                    # Allow if it looks like MFN
                    if not any(k in mt.lower() for k in ["third", "mfn", "103", "standard"]):
                        skipped += 1
                        continue

            code_raw = row.get(code_col, "") if code_col else ""
            code = normalize_code(code_raw)
            if not code or len(code) < 6:
                skipped += 1
                continue

            rate_raw = row.get(rate_col, "") if rate_col else ""
            rate = parse_rate(rate_raw)

            desc_raw = row.get(desc_col, "") if desc_col else ""
            desc = str(desc_raw).strip() if desc_raw else ""

            chapter = int(code[:2]) if code[:2].isdigit() else 0
            key = code[:8]  # CN8 level

            if key not in tariff or rate < tariff[key]["r"]:
                tariff[key] = {
                    "d": desc[:120],
                    "r": rate,
                    "ch": chapter,
                }
            processed += 1

            if processed % 5000 == 0:
                print(f"  Processed {processed} rows...")

        except Exception:
            skipped += 1
            continue

    return tariff, processed, skipped

def parse_excel(path: str) -> dict:
    """Parse TARIC Excel file (fallback for manual download)."""
    try:
        import openpyxl
    except ImportError:
        os.system(f"{sys.executable} -m pip install openpyxl --quiet")
        import openpyxl

    print(f"  Loading Excel: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row
    header = None
    header_row_idx = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        row_str = [str(c).lower() if c else "" for c in row]
        if any(k in " ".join(row_str) for k in ["nomenclature", "cn", "code", "tariff"]):
            header = list(row)
            header_row_idx = i
            break

    if not header:
        header = [f"col_{i}" for i in range(20)]
        header_row_idx = 1

    # Map columns
    col = {str(h).lower(): i for i, h in enumerate(header) if h}
    code_idx = next((col[k] for k in col if any(x in k for x in ["nomenclature","cn_code","tariff_code","commodity"])), 0)
    desc_idx = next((col[k] for k in col if "description" in k), 1)
    rate_idx = next((col[k] for k in col if any(x in k for x in ["duty","rate"])), 3)

    tariff = {}
    skipped = 0
    processed = 0

    for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
        try:
            code = normalize_code(row[code_idx] if code_idx < len(row) else None)
            if not code or len(code) < 6:
                skipped += 1
                continue
            rate = parse_rate(row[rate_idx] if rate_idx < len(row) else None)
            desc_raw = row[desc_idx] if desc_idx < len(row) else ""
            desc = str(desc_raw).strip() if desc_raw else ""
            chapter = int(code[:2]) if code[:2].isdigit() else 0
            key = code[:8]
            if key not in tariff:
                tariff[key] = {"d": desc[:120], "r": rate, "ch": chapter}
            processed += 1
        except Exception:
            skipped += 1
            continue

    wb.close()
    return tariff, processed, skipped

def build(force=False, excel_path=None):
    print("🇪🇺  EU TARIC Tariff Builder")
    print("─" * 40)

    current_version = check_version()
    stored_version  = load_stored_version()

    if not force and current_version == stored_version and os.path.exists(OUTPUT_PATH):
        print(f"✓  Already up to date (version: {current_version})")
        return False

    print(f"  Dataset version : {current_version}")
    print(f"  Stored version  : {stored_version or '(none)'}")

    tariff = {}
    processed = skipped = 0

    # ── Option A: Manual Excel path ────────────────────────────────────────────
    if excel_path:
        if not os.path.exists(excel_path):
            print(f"❌  Excel not found: {excel_path}")
            sys.exit(1)
        tariff, processed, skipped = parse_excel(excel_path)

    # ── Option B: Auto-download CSV ────────────────────────────────────────────
    else:
        print(f"\n  Attempting download from data.europa.eu...")
        try:
            req = urllib.request.Request(TARIC_CSV_URL, headers={"User-Agent": "iDuties/1.0"})
            with urllib.request.urlopen(req, timeout=120) as r:
                raw = r.read()
                # Handle zip if needed
                if raw[:4] == b"PK\x03\x04":
                    print("  Detected ZIP — extracting CSV...")
                    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                        csv_name = next((n for n in zf.namelist() if n.endswith(".csv")), None)
                        if csv_name:
                            raw = zf.read(csv_name)
                raw_text = raw.decode("utf-8-sig")
                print(f"  Downloaded {len(raw_text)//1024} KB")
                tariff, processed, skipped = parse_csv(raw_text)

        except urllib.error.URLError as e:
            print(f"\n⚠️   Auto-download failed: {e}")
            print("\n  The TARIC bulk download URL changes each year.")
            print("  Manual steps:")
            print("  1. Go to: https://taxation-customs.ec.europa.eu/customs/customs-tariff/eu-customs-tariff-taric_en")
            print("  2. Download the Excel file ('TARIC raw data in Excel format')")
            print("  3. Run: python builders/build_eu.py --excel path/to/taric.xlsx")
            print("\n  Or try data.europa.eu:")
            print("  https://data.europa.eu/data/datasets/eu-customs-tariff-taric")
            sys.exit(1)

    if not tariff:
        print("❌  No tariff data parsed. Check the source file format.")
        sys.exit(1)

    print(f"\n✓  Parsed {len(tariff)} unique CN codes ({processed} rows, {skipped} skipped)")

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(tariff, f, separators=(",", ":"))

    size_kb = os.path.getsize(OUTPUT_PATH) / 1024
    print(f"✓  Written to {OUTPUT_PATH}  ({size_kb:.0f} KB)")
    save_version(current_version)
    return True

if __name__ == "__main__":
    force = "--force" in sys.argv
    excel_path = None
    if "--excel" in sys.argv:
        idx = sys.argv.index("--excel")
        excel_path = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None
    build(force=force, excel_path=excel_path)
